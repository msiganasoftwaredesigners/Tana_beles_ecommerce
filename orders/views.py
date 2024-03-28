#orders/views.py
from django.shortcuts import render, redirect, HttpResponse
from .models import Order, OrderItem
from carts.models import Cart, CartItem

from .forms import OrderForm
from .export_to_gsheets import export_orders_to_gsheets
from django.http import HttpResponse
from openpyxl import Workbook
from reportlab.lib.pagesizes import letter, landscape
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.pagesizes import A3
from django.shortcuts import get_list_or_404
from carts.views import clear_cart

def checkout(request):
    orders = Order.objects.all()
    return render(request, 'orders.html', {'orders': orders})

def order_complete(request):
    return render(request, 'order_complete.html')


def export_orders_pdf(request):
    queryset = get_list_or_404(Order)
    return export_orders_view(request, queryset, 'pdf')

def export_orders_xlsx(request):
    queryset = get_list_or_404(Order)
    return export_orders_view(request, queryset, 'xlsx')

def export_orders(request):
    queryset = get_list_or_404(Order)
    export_orders_view(request, queryset, 'pdf')
    return export_orders_view(request, queryset, 'xlsx')

def export_orders_view(request, queryset, format):
    orders = queryset

    if format == 'xlsx':
        # Your existing Excel code...
        wb = Workbook()
        ws = wb.active

        headers = ['First Name', 'Last Name', 'Phone', 'Email', 'Date', 'Total Prices', 'Address', 'Payment Status', 'Status']
        for col_num, header in enumerate(headers, 1):
            col_letter = ws.cell(row=1, column=col_num).column_letter
            ws['{}1'.format(col_letter)] = header
            ws.column_dimensions[col_letter].width = 15

        for row_num, order in enumerate(orders, 2):
            ws.cell(row=row_num, column=1, value=order.first_name)
            ws.cell(row=row_num, column=2, value=order.last_name)
            ws.cell(row=row_num, column=3, value=order.order_phone)
            ws.cell(row=row_num, column=4, value=order.order_email)
            ws.cell(row=row_num, column=5, value=order.order_date.strftime('%Y-%m-%d'))
            ws.cell(row=row_num, column=6, value=str(order.order_total_prices))
            ws.cell(row=row_num, column=7, value=order.order_address)
            ws.cell(row=row_num, column=8, value=order.payment_status)
            ws.cell(row=row_num, column=9, value=order.status)

        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=orders.xlsx'
        wb.save(response)

    elif format == 'pdf':
        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = 'attachment; filename=orders.pdf'

        doc = SimpleDocTemplate(response, pagesize=landscape(A3))

        data = []
        headers = ['First Name', 'Last Name', 'Phone', 'Email', 'Date', 'Total Prices', 'Address', 'Payment Status', 'Status']
        data.append(headers)

        for order in orders:
            row = [order.first_name, order.last_name, order.order_phone, order.order_email, order.order_date.strftime('%Y-%m-%d'), str(order.order_total_prices), order.order_address, order.payment_status, order.status]
            data.append(row)

        table = Table(data)

        style = TableStyle([
            ('BACKGROUND', (0,0), (-1,0), colors.grey),
            ('TEXTCOLOR',(0,0),(-1,0),colors.whitesmoke),

            ('ALIGN',(0,0),(-1,-1),'CENTER'),
            ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
            ('FONTSIZE', (0,0), (-1,0), 14),

            ('BOTTOMPADDING', (0,0), (-1,0), 12),
            ('BACKGROUND',(0,1),(-1,-1),colors.beige),
            ('GRID',(0,0),(-1,-1),1,colors.black)
        ])
        table.setStyle(style)

        elements = []
        elements.append(table)

        doc.build(elements)
    else:
        return HttpResponse("Invalid format")

    return response

def create_order(request):
      # Get the current user's email
    user_email = request.user.email

    # Get the total price from the session
    total = request.session.get('total', 0)
    cart_id = request.session.get('cart_id')
    if request.method == 'POST':
        print(request.POST)
        form = OrderForm(request.POST)
        if form.is_valid():
            # Create a new Order object
            order = Order(
                first_name=form.cleaned_data['first_name'],
                last_name=form.cleaned_data['last_name'],
                order_email=user_email,
                order_address=form.cleaned_data['order_address'],
                referral_code=form.cleaned_data['referral_code'],
                order_total_prices=total,
                user = request.user
            )

            # Save the Order object to the database
            order.save()


            # Create OrderItem objects for each item in the cart
            cart = Cart.objects.get(cart_id=cart_id)
            cart_items = CartItem.objects.filter(cart=cart)
            for item in cart_items:
                OrderItem.objects.create(
                    order=order,
                    product=item.product,
                    product_name=item.product_name,
                    product_price=item.selected_price,
                    product_size=item.selected_size,
                    product_color=item.selected_color,
                    quantity=item.quantity,
                    product_brand=item.product_brand
                )

          
            # Redirect the user to a confirmation page
            clear_cart(request)
            return redirect('payment_page')
        else:
            print(form.errors)
    else:
        form = OrderForm(initial={'order_total_prices': total})

    return render(request, 'orders.html', {'form': form, 'total': total})


def export_view(request):
    export_orders_to_gsheets()
    return HttpResponse("Data exported to Google Sheets")